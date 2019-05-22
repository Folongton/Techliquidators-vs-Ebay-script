import os                   
import re                   
from openpyxl import load_workbook, Workbook 
from lot import Lot         
import requests             
from ebay import Ebay       
from consts import cat_url, directory, download_url 
from database import Database   
from helper import avg, median  
from bs4 import BeautifulSoup   


auction_names = {} 


def get_existed_items(db):
   ''' items = []
    for item in db.get('items', 'id'):
        items.append(int(item[0]))
    return items
    '''
   return [int(item[0]) for item in db.get('items', 'id')]   


def get_new_items():                            
    html = requests.get(cat_url).text          
    soup = BeautifulSoup(html, 'html.parser')   
    items = []                                  
    losts = soup.find_all('div', class_='truncate-ellipsis')
    for lot in losts:
        name = lot.span.a.text                   
        item = re.search('marketplace_main\.auction&amp;id=(\d+)"', str(lot)) 
        if item:                                 
            item = int(item.group(1))            
            auction_names.update({item: name})   
            print(auction_names)                
            items.append(item)
    return list(set(items))                     


def save_file(name, content):
    if not os.path.exists(directory):           
        os.makedirs(directory)                  
    if not os.path.exists(f'{directory}/{name}.xlsx'):
        open(f'{directory}/{name}.xlsx', 'wb').write(content)


if __name__ == '__main__':                   
                                             
    db = Database('techliquidators.sqlite')  

    existed_items = get_existed_items(db)
    new_items = get_new_items()
    new_items = [item for item in new_items if item not in existed_items]   

    # temp = []
    # for item in new_items:
    #     if item not in existed_items:
    #         temp.append(item)

    db.insert_many('items', new_items)      
    db.close()

    ebay = Ebay('VasylZhe-TechPric-PRD-2393f3dea-5cde0afe')     

    for itemID in new_items:
        if not os.path.exists(f'{directory}/{itemID}.xlsx'):
            r = requests.post(download_url, {'auctionID': itemID})  
            save_file(itemID, r.content)                           

    for file in os.listdir(directory):    
        wb = Workbook()                   
        ws = wb.active                    
        ws.append(
            ['Auction Name', 'Auction ID', 'MFG Name', 'MFG Part Number', 'Title', 'UPC',
             'N', 'Name AVG price by name EBay', 'Name Median  price by name Ebay using high to low sorting',
             'N', 'MFG AVG price by MFG + Part Number Ebay', 'MFG Median (Median price by MFG + Part Number Ebay using hight to low sorting)',
             'N', 'UPC AVG price by UPC Ebay', 'UPC Median price by UPC Ebay using  hight to low sorting', ' '])    # appends top row for more details here: https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

        id_ = int(file.replace('.xlsx', ''))                       
        lots = []                                              
        temp_wb = load_workbook(f'{directory}/{file}')           
        temp_ws = temp_wb.active                               
        for row in range(2, temp_ws.max_row + 1):                
            values = []                                            
            for column in range(1, temp_ws.max_column + 1):        
                cell_obj = temp_ws.cell(row=row, column=column)    
                values.append(cell_obj.value)                      
            lot = Lot(List=values)
            lot.fill_listings_by_Title(ebay)
            lot.fill_listings_by_MFG(ebay)
            lot.fill_listings_by_UPC(ebay)
            lots.append(lot)
            ws.append([auction_names.get(id_), lot.AuctionID,
                       lot.MFGName, lot.MFGPartNumber, lot.Title, lot.UPC,
                       len(lot.listingsByTitle), avg(lot.listingsByTitle), median(lot.listingsByTitle),
                       len(lot.listingsByMFG), avg(lot.listingsByMFG), median(lot.listingsByMFG),
                       len(lot.listingsByUPC), avg(lot.listingsByUPC), median(lot.listingsByUPC)])
        index = ws.max_row + 1
        ws[f'H{index}'] = sum([avg(l.listingsByTitle) for l in lots])
        ws[f'I{index}'] = sum([median(l.listingsByTitle) for l in lots])
        ws[f'K{index}'] = sum([avg(l.listingsByMFG) for l in lots])
        ws[f'L{index}'] = sum([median(l.listingsByMFG) for l in lots])
        ws[f'N{index}'] = sum([avg(l.listingsByUPC) for l in lots])
        ws[f'O{index}'] = sum([median(l.listingsByUPC) for l in lots])

        index = ws.max_row + 2
        total_avg_upc = (sum([avg(l.listingsByUPC) for l in lots]) + sum([median(l.listingsByUPC) for l in lots]))/2
        ws[f'N{index}'] = '30% ROI:'
        ws[f'O{index}'] = total_avg_upc * 0.77
        ws[f'P{index}'] = '42% ROI:'
        ws[f'Q{index}'] = total_avg_upc * 0.7

        index = ws.max_row + 1
        ws[f'N{index}'] = 'Ebay 0.13 fee'
        ws[f'O{index}'] = total_avg_upc * 0.13
        ws[f'Q{index}'] = total_avg_upc * 0.13

        index = ws.max_row + 1
        ws[f'N{index}'] = 'Shipping'
        ws[f'O{index}'] = '200'
        ws[f'Q{index}'] = '200'

        index = ws.max_row + 1
        ws[f'N{index}'] = 'Earnings'
        ws[f'O{index}'] = total_avg_upc - (total_avg_upc * 0.77) - (total_avg_upc * 0.13) - 200
        ws[f'Q{index}'] = total_avg_upc - (total_avg_upc * 0.7) - (total_avg_upc * 0.13) - 200

        index = ws.max_row + 1
        ws[f'N{index}'] = 'ROI clean'
        ws[f'O{index}'] = ((total_avg_upc - (total_avg_upc * 0.77) - (total_avg_upc * 0.13) - 200) / total_avg_upc)*100
        ws[f'Q{index}'] = ((total_avg_upc - (total_avg_upc * 0.7) - (total_avg_upc * 0.13) - 200) / total_avg_upc)*100

        # os.remove(f'{directory}/{file}')
        print('save EXCEL file')
        wb.save(f'{directory}/{file}')

        os.chdir(r'D:\Study 2018\Python codes 2\Techliquidators script\files')
        saved_file = [f for f in os.listdir('.') if os.path.isfile(f)]
        for f in saved_file:
            os.rename(f, f'D:/Study 2018/Python codes 2/Techliquidators script/UNPROCESSED auctions/{f}')
            print(f'file(s): {f} moved to Folder: UNPROCESSED auctions')





   
    # temp_wb = load_workbook(BytesIO(r.content))
    # temp_ws = temp_wb.active
    # for row in range(temp_ws.max_row + 1):
    #       print(temp_ws['A1'].value)
